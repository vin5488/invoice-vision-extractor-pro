# -*- coding: utf-8 -*-
"""Test.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/13Qj5UDXHoF5gQIeJA9vaS6RIwPJo4k_A
"""

import os
import cv2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font  # For header styling
import argparse
from typing import Dict, Any, Optional
import json

# For interactive file dialogs using Tkinter
try:
    import tkinter.filedialog
    from tkinter import Tk
except ImportError:
    Tk = None


class InvoiceExporter:
    """
    A class to export invoice data to Excel files in different formats,
    including a customizable export naming.
    """

    def __init__(self):
        """Initialize the InvoiceExporter"""
        pass

    @staticmethod
    def get_default_manufacturing_template_data() -> Dict[str, Any]:
        """Generate default manufacturing template data"""
        return {
            "stateName": "Karnataka, Code : 29",
            "termsOfDelivery": "As per terms",
            "items": [
                {
                    "partNo": "Laser Cutting-MIT-EA012B014-04",
                    "description": "SIZE:147.4X179.7X6MM-CUT LENGTH-1207MM-HR",
                    "hsn": "73269070",
                    "quantity": "116 Nos.",
                    "rate": "118.500",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "13,746.000"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA015C294-02",
                    "description": "SIZE:110X222.4X6MM-CUT LENGTH-949MM-HR",
                    "hsn": "73269070",
                    "quantity": "100 Nos.",
                    "rate": "103.000",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "10,300.000"
                },
                # Additional items omitted for brevity...
            ]
        }

    def create_manufacturing_excel(self, data: Optional[Dict[str, Any]] = None,
                                   output_dir: str = ".", export_name: str = "Invoice_Export") -> str:
        """
        Create a manufacturing invoice Excel file.

        If the provided data represents multiple invoices (a list inside the key 'manufacturingTableData'),
        the method aggregates the items into a unified invoice.

        Args:
            data: Dictionary containing invoice data.
            output_dir: Directory in which to save the Excel file.
            export_name: Custom filename prefix for the exported Excel file.

        Returns:
            Path to the created Excel file.
        """
        # Use default data if none is provided or if data is empty
        if data is None or not data:
            manufacturing_template = self.get_default_manufacturing_template_data()
        else:
            manufacturing_template = data.get('manufacturingTableData', self.get_default_manufacturing_template_data())

        # If manufacturing_template is a list, aggregate the items.
        if isinstance(manufacturing_template, list):
            aggregated_items = []
            for invoice in manufacturing_template:
                if isinstance(invoice.get("items"), list):
                    for item in invoice["items"]:
                        if "partNo" not in item:
                            aggregated_items.append({
                                "partNo": invoice.get("invoiceNumber", ""),
                                "description": item.get("description", ""),
                                "hsn": "",
                                "quantity": item.get("quantity", ""),
                                "rate": item.get("unitPrice", ""),
                                "per": "Nos.",
                                "discountPercentage": "",
                                "amount": item.get("total", "")
                            })
                        else:
                            aggregated_items.append(item)
                else:
                    aggregated_items.append({
                        "partNo": invoice.get("invoiceNumber", ""),
                        "description": "Invoice Total",
                        "hsn": "",
                        "quantity": "",
                        "rate": "",
                        "per": "",
                        "discountPercentage": "",
                        "amount": invoice.get("total", "")
                    })
            manufacturing_template = {
                "stateName": "Aggregated Invoices",
                "termsOfDelivery": "",
                "items": aggregated_items
            }

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Manufacturing Invoice"

        # Add header information
        ws['A1'] = "State Name"
        ws['B1'] = ":"
        ws['C1'] = manufacturing_template.get("stateName", "")
        ws['G1'] = "Terms of Delivery"
        ws['H1'] = manufacturing_template.get("termsOfDelivery", "")

        # Table headers on row 3
        headers = ["SI No.", "Part No", "Description of Goods", "HSN/SAC",
                   "Quantity", "Rate", "per", "Disc. %", "Amount"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)

        # Add invoice items starting at row 4
        items = manufacturing_template.get("items", [])
        for row_idx, item in enumerate(items, 4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3)
            ws.cell(row=row_idx, column=2, value=item.get("partNo", ""))
            ws.cell(row=row_idx, column=3, value=item.get("description", ""))
            ws.cell(row=row_idx, column=4, value=item.get("hsn", ""))
            ws.cell(row=row_idx, column=5, value=item.get("quantity", ""))
            ws.cell(row=row_idx, column=6, value=item.get("rate", ""))
            ws.cell(row=row_idx, column=7, value=item.get("per", ""))
            ws.cell(row=row_idx, column=8, value=item.get("discountPercentage", ""))
            ws.cell(row=row_idx, column=9, value=item.get("amount", ""))

        # Set column widths for readability
        column_widths = [5, 25, 30, 10, 10, 10, 5, 10, 15]
        for col_idx, width in enumerate(column_widths, 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = width

        # Build and save the Excel file using export_name and current timestamp
        timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"{export_name}_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        wb.save(output_path)
        print(f"Excel file saved as {output_path}")

        return output_path

    def extract_data_from_image(self, image_path: str) -> Dict[str, Any]:
        """
        Extract invoice data from an image using OpenCV.

        Args:
            image_path: Path to the invoice image.

        Returns:
            Dictionary containing extracted invoice data.
        """
        print(f"Processing image: {image_path}")

        # Load image using OpenCV
        image = cv2.imread(image_path)
        if image is None:
            raise ValueError(f"Could not read image file: {image_path}")

        # Preprocess image: convert to grayscale and apply thresholding for OCR
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        _, threshold = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

        # Placeholder for OCR functionality (integrate pytesseract as needed)
        print("Note: This is a placeholder for OCR functionality. Integrate pytesseract for real OCR.")

        # Return sample/mock data
        return {
            "invoiceNumber": f"INV-{os.path.basename(image_path).split('.')[0]}",
            "date": pd.Timestamp.now().strftime('%Y-%m-%d'),
            "materialId": f"MAT-{pd.Timestamp.now().strftime('%d%H%M')}",
            "items": [
                {
                    "description": "Laser Cut Item",
                    "quantity": 10,
                    "unitPrice": "100.00",
                    "total": "1000.00"
                }
            ]
        }


def process_file(file_path: str, exporter: InvoiceExporter) -> Dict[str, Any]:
    """
    Process a file (image, PDF, or JSON) and extract invoice data.

    Args:
        file_path: Path to the file.
        exporter: InvoiceExporter instance.

    Returns:
        Dictionary containing extracted data.
    """
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    if ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
        return exporter.extract_data_from_image(file_path)
    elif ext == '.pdf':
        print("PDF processing not implemented yet. Returning placeholder data.")
        return {
            "invoiceNumber": f"INV-{os.path.basename(file_path).split('.')[0]}",
            "date": pd.Timestamp.now().strftime('%Y-%m-%d'),
            "total": "$1,234.56"
        }
    elif ext in ['.json']:
        with open(file_path, 'r') as f:
            return json.load(f)
    else:
        print(f"Unsupported file type: {ext}")
        return {}


def main():
    parser = argparse.ArgumentParser(description='Invoice OCR and Export Tool')
    parser.add_argument('--input', '-i',
                        help='Input file path or use "upload" to select interactively', default='.')
    parser.add_argument('--output', '-o',
                        help='Output directory or use "choose" to select interactively', default='.')
    parser.add_argument('--export-only',
                        action='store_true', help='Skip OCR and only export sample data')
    parser.add_argument('--export-name',
                        help='Custom filename prefix for the exported Excel file (without extension)',
                        default='Invoice_Export')
    # Use parse_known_args to avoid issues with extra arguments in some environments.
    args, _ = parser.parse_known_args()

    # Handle interactive file selection for input if needed.
    input_path = args.input
    if input_path.lower() == "upload":
        if Tk is None:
            print("Tkinter is not available, cannot open file dialog.")
            return
        root = Tk()
        root.withdraw()  # Hide the main window
        file_path = tkinter.filedialog.askopenfilename(title="Select file to process")
        if file_path:
            input_path = file_path
            print(f"File selected: {input_path}")
        else:
            print("No file selected. Exiting.")
            return

    # Handle interactive directory selection for output if requested.
    output_dir = args.output
    if output_dir.lower() == "choose":
        if Tk is None:
            print("Tkinter is not available, cannot open directory chooser.")
            return
        root = Tk()
        root.withdraw()
        dir_path = tkinter.filedialog.askdirectory(title="Select output directory")
        if dir_path:
            output_dir = dir_path
            print(f"Output directory selected: {output_dir}")
        else:
            print("No directory selected. Exiting.")
            return

    exporter = InvoiceExporter()

    # If export-only flag is set, export sample data and exit.
    if args.export_only:
        print("Exporting sample invoice...")
        output_file = exporter.create_manufacturing_excel(None,
                                                          output_dir=output_dir,
                                                          export_name=args.export_name)
        print(f"Sample invoice exported to: {output_file}")
        return

    # Process the input. If input_path is a directory, process all files inside;
    # if it is a file, just process that one.
    all_data = []
    if os.path.isdir(input_path):
        for file_name in os.listdir(input_path):
            file_path = os.path.join(input_path, file_name)
            if os.path.isfile(file_path):
                try:
                    data = process_file(file_path, exporter)
                    if data:
                        all_data.append(data)
                except Exception as e:
                    print(f"Error processing {file_path}: {e}")
    elif os.path.isfile(input_path):
        try:
            data = process_file(input_path, exporter)
            if data:
                all_data.append(data)
        except Exception as e:
            print(f"Error processing {input_path}: {e}")
    else:
        print("The input path is neither a file nor a directory. Exiting.")
        return

    if all_data:
        # If multiple files produced data, aggregate and export to Excel.
        output_file = exporter.create_manufacturing_excel(
            {"manufacturingTableData": all_data},
            output_dir=output_dir,
            export_name=args.export_name)
        print(f"Data exported to: {output_file}")
    else:
        print("No data to export.")


if __name__ == "__main__":
    main()















import os
import cv2
import numpy as np
import pytesseract
import pandas as pd
from pdf2image import convert_from_path

# If tesseract is not in your PATH, set it here:
# pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def process_file(input_path, output_excel='invoice_output.xlsx', poppler_path=None):
    """
    Main function to handle both PDF and image files.
    1) If PDF, convert each page to an image and process them.
    2) If image, just process that single file.
    3) Combine results into a single Excel with two sheets:
       - FullText
       - ParsedTable
    """

    all_full_text = []    # List of strings, one per page
    all_tables = []       # List of DataFrames, one per page

    # 1) Check if it's a PDF
    if input_path.lower().endswith('.pdf'):
        pages = convert_from_path(input_path, dpi=200, poppler_path=poppler_path)
        for i, page in enumerate(pages):
            temp_img = f'temp_page_{i}.png'
            page.save(temp_img, 'PNG')

            page_text, page_table = process_image(temp_img)
            all_full_text.append(f"--- Page {i+1} ---\n" + page_text)

            # Add a "Page" column to distinguish which page
            if not page_table.empty:
                page_table.insert(0, 'Page', i+1)
            all_tables.append(page_table)

            # Cleanup temp image
            os.remove(temp_img)

    else:
        # 2) It's an image (PNG/JPG/etc.)
        page_text, page_table = process_image(input_path)
        all_full_text.append(page_text)
        if not page_table.empty:
            page_table.insert(0, 'Page', 1)
        all_tables.append(page_table)

    # Combine all text from pages into one
    combined_text = "\n\n".join(all_full_text)

    # Combine all table DataFrames (if multiple pages)
    if len(all_tables) > 0:
        combined_table = pd.concat(all_tables, ignore_index=True)
    else:
        combined_table = pd.DataFrame()

    # 3) Write to Excel (two sheets)
    with pd.ExcelWriter(output_excel) as writer:
        # Full text in one cell
        df_text = pd.DataFrame([combined_text], columns=['Full Text'])
        df_text.to_excel(writer, sheet_name='FullText', index=False)

        # Table in second sheet
        if not combined_table.empty:
            combined_table.to_excel(writer, sheet_name='ParsedTable', index=False)

    print(f"✅ Finished! Results saved to: {output_excel}")


def process_image(img_path):
    """
    Processes a single image:
    1) Full-page OCR for all text
    2) Table detection with morphological operations + contour
    3) OCR each cell -> DataFrame
    Returns: (full_text_string, table_df)
    """

    # --- (A) Full Page OCR ---
    image = cv2.imread(img_path)
    config = r'--oem 3 --psm 6'
    full_text = pytesseract.image_to_string(image, config=config)

    # --- (B) Table Detection ---
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # Invert the bits (for easier line detection)
    # adaptiveThreshold with ~gray in some tutorials, or we can do ~gray if needed
    binary = cv2.adaptiveThreshold(~gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                   cv2.THRESH_BINARY, 15, -2)

    # Extract horizontal and vertical lines
    horizontal = binary.copy()
    vertical = binary.copy()

    scale = 20  # Adjust if your invoice is very large or small
    horizontalsize = horizontal.shape[1] // scale
    verticalsize = vertical.shape[0] // scale

    # Morph for horizontal lines
    horizontalStructure = cv2.getStructuringElement(cv2.MORPH_RECT, (horizontalsize, 1))
    horizontal = cv2.erode(horizontal, horizontalStructure)
    horizontal = cv2.dilate(horizontal, horizontalStructure)

    # Morph for vertical lines
    verticalStructure = cv2.getStructuringElement(cv2.MORPH_RECT, (1, verticalsize))
    vertical = cv2.erode(vertical, verticalStructure)
    vertical = cv2.dilate(vertical, verticalStructure)

    # Combine them
    mask = horizontal + vertical

    # Find contours
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    boxes = []
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        # Filter out too small boxes (noise)
        if w > 40 and h > 20:
            boxes.append((x, y, w, h))

    # Sort boxes top-to-bottom, then left-to-right
    boxes = sorted(boxes, key=lambda b: (b[1], b[0]))

    # --- (C) Group boxes into rows ---
    row_threshold = 10
    rows = []
    current_row = []
    last_y = -100

    for (x, y, w, h) in boxes:
        if abs(y - last_y) > row_threshold:
            # start a new row
            if current_row:
                # sort left-to-right
                current_row = sorted(current_row, key=lambda b: b[0])
                rows.append(current_row)
            current_row = [(x, y, w, h)]
            last_y = y
        else:
            current_row.append((x, y, w, h))

    # Add the final row
    if current_row:
        current_row = sorted(current_row, key=lambda b: b[0])
        rows.append(current_row)

    # --- (D) OCR each cell in each row ---
    table_data = []
    for row in rows:
        row_data = []
        for (x, y, w, h) in row:
            roi = gray[y:y+h, x:x+w]
            cell_text = pytesseract.image_to_string(roi, config=config)
            cell_text = cell_text.strip().replace('\n',' ')
            row_data.append(cell_text)
        table_data.append(row_data)

    # If no table_data found, return empty DataFrame
    if not table_data:
        return (full_text, pd.DataFrame())

    # Some invoices might have 8-9 columns, some 6, etc.
    # We'll guess the maximum columns from the largest row
    max_cols = max(len(r) for r in table_data)
    # We'll create generic column names
    col_names = [f'Col_{i+1}' for i in range(max_cols)]

    df_table = pd.DataFrame(table_data, columns=col_names)

    # Optionally rename columns if you know the exact structure:
    # Example (if you know you have exactly 8 columns):
    # desired_cols = ['Sl No','Part No','Description','HSN/SAC','Quantity','Rate','Disc.%','Amount']
    # if max_cols >= 8:
    #     df_table.columns = desired_cols + col_names[8:]  # keep extras

    return (full_text, df_table)


if __name__ == "__main__":
    # EXAMPLE USAGE
    # -------------
    # 1) For an image (PNG/JPG)
    # process_file("invoice_1.png", "output_invoice.xlsx")
    #
    # 2) For a PDF with multiple pages
    # process_file("sample_invoice.pdf", "output_invoice.xlsx", poppler_path=r"C:\poppler-23.05.0\Library\bin")

    # Adjust the path below to your file
    input_file = "invoice_1.png"
    output_file = "invoice_output.xlsx"
    process_file(input_file, output_file)