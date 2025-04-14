
import os
import cv2
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
import argparse
from typing import Dict, List, Any, Optional
import json

class InvoiceExporter:
    """
    A class to export invoice data to Excel files in different formats,
    including laser cutting invoice format.
    """

    def __init__(self):
        """Initialize the InvoiceExporter"""
        # Try to import pytesseract if available
        try:
            import pytesseract
            self.pytesseract = pytesseract
            self.ocr_available = True
        except ImportError:
            print("Warning: pytesseract not found. OCR functionality will be limited.")
            self.ocr_available = False

    @staticmethod
    def get_default_manufacturing_template_data() -> Dict[str, Any]:
        """Generate default manufacturing template data"""
        # ... keep existing code (default template data)
        return {
            "stateName": "Karnataka, Code : 29",
            "termsOfDelivery": "As per terms",
            "items": [
                {
                    "partNo": "Laser Cutting-MIT-EA012B014-04",
                    "description": "SIZE:147.4X179.7X6MM-CUT LENGTH-1207MM-HR",
                    "hsn": "73269070",
                    "quantity": "116 Nos.",
                    "rate": "118.500",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "13,746.000"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA015C294-02",
                    "description": "SIZE:110X222.4X6MM-CUT LENGTH-949MM-HR",
                    "hsn": "73269070",
                    "quantity": "100 Nos.",
                    "rate": "103.000",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "10,300.000"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA021C281-05",
                    "description": "SIZE:110X125X6MM-CUT LENGTH-543MM-HR",
                    "hsn": "73269070",
                    "quantity": "10 Nos.",
                    "rate": "58.400",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "584.000"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA033C300-01",
                    "description": "SIZE:170X240X6MM-CUT LENGTH-1287MM-HR",
                    "hsn": "73269070",
                    "quantity": "20 Nos.",
                    "rate": "160.900",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "3,218.000"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA033C301-01",
                    "description": "SIZE:125X145X6MM-CUT LENGTH-768MM-HR",
                    "hsn": "73269070",
                    "quantity": "10 Nos.",
                    "rate": "78.900",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "789.000"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA111B538-04 REV 0",
                    "description": "SIZE:175X355X6MM- CUT LENGTH1384MM-HR",
                    "hsn": "73269070",
                    "quantity": "12 Nos.",
                    "rate": "223.300",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "2,679.600"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA131D685-01",
                    "description": "SIZE:32X32X6MM-CUT LENGTH -157MM-HR 2062",
                    "hsn": "73269070",
                    "quantity": "180 Nos.",
                    "rate": "8.800",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "1,584.000"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA175C796-01",
                    "description": "SIZE:115X255.4X6MM-CUT LENGTH-1322MM-HR",
                    "hsn": "73269070",
                    "quantity": "28 Nos.",
                    "rate": "130.800",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "3,662.400"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA179D554-01",
                    "description": "SIZE:100X110X6MM-CUT LENGTH-664MM-HR",
                    "hsn": "73269070",
                    "quantity": "10 Nos.",
                    "rate": "55.900",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "559.000"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA214C825-01LS1",
                    "description": "SIZE:50X1155X6MM-CUT LENGTH 2617MM-HR",
                    "hsn": "73269070",
                    "quantity": "6 Nos.",
                    "rate": "257.900",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "1,547.400"
                },
                {
                    "partNo": "Laser Cutting-MIT-EA214C825-01LS2",
                    "description": "SIZE:50X1230X6MM-CUT LENGTH 2767MM-HR",
                    "hsn": "73269070",
                    "quantity": "6 Nos.",
                    "rate": "273.900",
                    "per": "Nos.",
                    "discountPercentage": "",
                    "amount": "1,643.400"
                }
            ]
        }

    def create_manufacturing_excel(self, data: Optional[Dict[str, Any]] = None) -> str:
        """
        Create a manufacturing invoice Excel file
        
        Args:
            data: Dictionary containing invoice data
            
        Returns:
            Path to the created Excel file
        """
        # ... keep existing code (Excel creation functionality)
        # Use default data if none provided
        if data is None or not data:
            manufacturing_template = self.get_default_manufacturing_template_data()
        else:
            manufacturing_template = data.get('manufacturingTableData', self.get_default_manufacturing_template_data())
        
        # Create workbook and active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Manufacturing Invoice"
        
        # Add header information
        ws['A1'] = "State Name"
        ws['B1'] = ":"
        ws['C1'] = manufacturing_template.get("stateName", "")
        ws['G1'] = "Terms of Delivery"
        ws['H1'] = manufacturing_template.get("termsOfDelivery", "")
        
        # Add table headers - row 3 (row 2 is empty)
        headers = ["SI No.", "Part No", "Description of Goods", "HSN/SAC", 
                   "Quantity", "Rate", "per", "Disc. %", "Amount"]
        
        for col_idx, header in enumerate(headers, 1):  # Start at column 1 (A)
            ws.cell(row=3, column=col_idx, value=header).font = Font(bold=True)
        
        # Add items data
        items = manufacturing_template.get("items", [])
        for row_idx, item in enumerate(items, 4):  # Start at row 4
            ws.cell(row=row_idx, column=1, value=row_idx - 3)  # SI No.
            ws.cell(row=row_idx, column=2, value=item.get("partNo", ""))
            ws.cell(row=row_idx, column=3, value=item.get("description", ""))
            ws.cell(row=row_idx, column=4, value=item.get("hsn", ""))
            ws.cell(row=row_idx, column=5, value=item.get("quantity", ""))
            ws.cell(row=row_idx, column=6, value=item.get("rate", ""))
            ws.cell(row=row_idx, column=7, value=item.get("per", ""))
            ws.cell(row=row_idx, column=8, value=item.get("discountPercentage", ""))
            ws.cell(row=row_idx, column=9, value=item.get("amount", ""))
            
        # Set column widths
        column_widths = [5, 25, 30, 10, 10, 10, 5, 10, 15]
        for col_idx, width in enumerate(column_widths, 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = width
        
        # Save workbook
        output_filename = f"Invoice_Export_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(output_filename)
        print(f"Excel file saved as {output_filename}")
        
        return output_filename

    def extract_data_from_image(self, image_path: str) -> Dict[str, Any]:
        """
        Extract invoice data from an image using OpenCV and OCR
        
        Args:
            image_path: Path to the invoice image
            
        Returns:
            Dictionary containing extracted invoice data
        """
        print(f"Processing image: {image_path}")
        
        # Load image
        image = cv2.imread(image_path)
        if image is None:
            raise ValueError(f"Could not read image file: {image_path}")

        # Enhanced table detection and extraction
        extracted_data = self.detect_and_extract_table(image)
        
        # If extraction failed or no tables found, return placeholder data
        if not extracted_data or not extracted_data.get("items"):
            print("Warning: Could not extract tabular data properly, returning placeholder data")
            return {
                "invoiceNumber": f"INV-{os.path.basename(image_path).split('.')[0]}",
                "date": pd.Timestamp.now().strftime('%Y-%m-%d'),
                "materialId": f"MAT-{pd.Timestamp.now().strftime('%d%H%M')}",
                "items": [
                    {
                        "description": "Extracted Item",
                        "quantity": "1",
                        "unitPrice": "100.00",
                        "total": "100.00"
                    }
                ]
            }
            
        return extracted_data

    def detect_and_extract_table(self, image: np.ndarray) -> Dict[str, Any]:
        """
        Enhanced table detection and cell extraction from invoice image
        
        Args:
            image: OpenCV image object
            
        Returns:
            Dictionary containing extracted table data
        """
        # Convert to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # Apply adaptive thresholding
        thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                      cv2.THRESH_BINARY_INV, 11, 2)
        
        # Dilate to connect text in cells
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
        dilated = cv2.dilate(thresh, kernel, iterations=1)
        
        # Find contours
        contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # Filter contours to get table cells (eliminate too small contours)
        min_cell_area = image.shape[0] * image.shape[1] / 500  # Adaptive threshold based on image size
        cells = []
        
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            area = w * h
            
            # Filter out noise and keep only cell-like rectangles
            if area > min_cell_area and w > 30 and h > 15:
                # Add padding to make sure we capture entire cell content
                padding = 2
                x_with_padding = max(0, x - padding)
                y_with_padding = max(0, y - padding)
                w_with_padding = min(image.shape[1] - x_with_padding, w + 2*padding)
                h_with_padding = min(image.shape[0] - y_with_padding, h + 2*padding)
                
                cells.append((x_with_padding, y_with_padding, w_with_padding, h_with_padding))
        
        # Sort cells by y-coordinate to group rows
        cells.sort(key=lambda c: c[1])
        
        # Group cells into rows based on y-position
        y_tolerance = image.shape[0] / 30  # Adaptive tolerance based on image height
        rows = []
        current_row = []
        
        if cells:
            current_row = [cells[0]]
            prev_y = cells[0][1]
            
            for i in range(1, len(cells)):
                x, y, w, h = cells[i]
                
                # If this cell is on a new row
                if abs(y - prev_y) > y_tolerance:
                    # Sort current row by x-coordinate
                    current_row.sort(key=lambda c: c[0])
                    rows.append(current_row)
                    current_row = [cells[i]]
                    prev_y = y
                else:
                    current_row.append(cells[i])
            
            # Add the last row
            if current_row:
                current_row.sort(key=lambda c: c[0])
                rows.append(current_row)
        
        # Parse table content using OCR if available
        items = []
        headers = []
        
        # Extract text from each cell
        for i, row in enumerate(rows):
            row_data = []
            
            for j, (x, y, w, h) in enumerate(row):
                # Extract cell ROI
                cell_image = gray[y:y+h, x:x+w]
                
                # Preprocess cell for better OCR
                _, cell_binary = cv2.threshold(cell_image, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
                
                # Apply OCR to the cell
                if self.ocr_available:
                    text = self.pytesseract.image_to_string(cell_binary).strip()
                else:
                    # If OCR isn't available, use cell position as placeholder
                    text = f"Cell_{i}_{j}"
                
                row_data.append(text)
            
            if i == 0:  # First row is likely headers
                headers = row_data
            else:  # Other rows are data
                # Create item dictionary based on likely column meanings
                item = {}
                
                # Try to match data with common invoice fields
                for j, cell_text in enumerate(row_data):
                    col_name = f"col_{j}" if j >= len(headers) else headers[j]
                    
                    # Map to standard fields based on typical invoice structure
                    # This is a simplistic approach - real mapping would be more complex
                    if j == 0:  # First column often contains part numbers or IDs
                        item["partNo"] = cell_text
                    elif j == 1:  # Second column typically has descriptions
                        item["description"] = cell_text
                    elif j == 2:  # Third column might be HSN code
                        item["hsn"] = cell_text
                    elif j == 3:  # Fourth column often has quantity
                        item["quantity"] = cell_text
                    elif j == 4:  # Fifth column might have rates/prices
                        item["rate"] = cell_text
                    elif j == 5:  # Sixth column might have unit (per)
                        item["per"] = cell_text
                    elif j == 6:  # Seventh column might have discount
                        item["discountPercentage"] = cell_text
                    elif j == 7:  # Eighth column might have amount/total
                        item["amount"] = cell_text
                    else:
                        item[f"field_{j}"] = cell_text
                
                if any(item.values()):  # Only add items that have some data
                    items.append(item)
        
        # Construct and return the extracted data
        return {
            "invoiceNumber": f"INV-{pd.Timestamp.now().strftime('%Y%m%d%H%M%S')}",
            "date": pd.Timestamp.now().strftime('%Y-%m-%d'),
            "stateName": "Extracted from Image",
            "termsOfDelivery": "Standard",
            "items": items
        }

def process_file(file_path: str, exporter: InvoiceExporter) -> Dict[str, Any]:
    """
    Process a file (image or PDF) and extract invoice data
    
    Args:
        file_path: Path to the file
        exporter: InvoiceExporter instance
        
    Returns:
        Dictionary containing extracted data
    """
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    if ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
        return exporter.extract_data_from_image(file_path)
    elif ext == '.pdf':
        print("PDF processing not implemented yet. Returning placeholder data.")
        return {
            "invoiceNumber": f"INV-{os.path.basename(file_path).split('.')[0]}",
            "date": pd.Timestamp.now().strftime('%Y-%m-%d'),
            "total": "$1,234.56"
        }
    elif ext in ['.json']:
        with open(file_path, 'r') as f:
            return json.load(f)
    else:
        print(f"Unsupported file type: {ext}")
        return {}

def main():
    parser = argparse.ArgumentParser(description='Invoice OCR and Export Tool')
    parser.add_argument('--input', '-i', help='Input directory or file path', default='.')
    parser.add_argument('--output', '-o', help='Output directory', default='.')
    parser.add_argument('--export-only', action='store_true', help='Skip OCR and only export sample data')
    args = parser.parse_args()
    
    exporter = InvoiceExporter()
    
    if args.export_only:
        print("Exporting sample invoice...")
        output_file = exporter.create_manufacturing_excel(None)
        print(f"Sample invoice exported to: {output_file}")
        return
    
    # Process input files
    input_path = args.input
    all_data = []
    
    if os.path.isdir(input_path):
        for file_name in os.listdir(input_path):
            file_path = os.path.join(input_path, file_name)
            if os.path.isfile(file_path):
                try:
                    data = process_file(file_path, exporter)
                    if data:
                        all_data.append(data)
                except Exception as e:
                    print(f"Error processing {file_path}: {e}")
    elif os.path.isfile(input_path):
        try:
            data = process_file(input_path, exporter)
            if data:
                all_data.append(data)
        except Exception as e:
            print(f"Error processing {input_path}: {e}")
    
    if all_data:
        # Export data to Excel
        output_file = exporter.create_manufacturing_excel({"manufacturingTableData": all_data})
        print(f"Data exported to: {output_file}")
    else:
        print("No data to export.")

if __name__ == "__main__":
    main()
